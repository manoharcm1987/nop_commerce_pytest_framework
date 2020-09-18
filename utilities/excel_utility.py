import openpyxl
from utilities.read_config import ConfigParser


def get_row_count(file_name, sheet):
    workbook = openpyxl.load_workbook(file_name)
    working_sheet = workbook[sheet]
    return working_sheet.max_row


def get_column_count(file_name, sheet):
    workbook = openpyxl.load_workbook(file_name)
    working_sheet = workbook[sheet]
    return  working_sheet.max_column


def read_cell_data(file_name, sheet, row_no, column_no):
    workbook = openpyxl.load_workbook(file_name)
    working_sheet = workbook[sheet]
    return working_sheet.cell(row=row_no, column=column_no).value


def write_data_to_cell(file_name, sheet, row_no, column_no, data):
    workbook = openpyxl.load_workbook(file_name)
    working_sheet = workbook[sheet]
    working_sheet.cell(row=row_no, column=column_no).value = data
    workbook.save(file_name)


def load_excel_data(path, sheet):
    workbook = openpyxl.load_workbook(path)
    working_sheet = workbook[sheet]
    data_list = list()
    for row in range(2, working_sheet.max_row+1):
        user_name = working_sheet.cell(row=row, column=1).value
        password = working_sheet.cell(row=row, column=2).value
        expected = working_sheet.cell(row=row, column=3).value
        user_data = (user_name, password, expected)
        data_list.append(user_data)
    return data_list
